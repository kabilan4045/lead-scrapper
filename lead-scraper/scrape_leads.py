#!/usr/bin/env python3
"""Scrape local business leads from Google Maps (via Apify) and append new,
deduplicated rows into the "Leads" tab of an existing Excel lead tracker."""

import argparse
import os
import re
import sys
import time
from datetime import date

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

APIFY_ACTOR_ID = "lukaskrivka~google-maps-with-contact-details"
APIFY_BASE_URL = "https://api.apify.com/v2"
POLL_INTERVAL_SECONDS = 5
MAX_POLL_SECONDS = 900

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_EXCEL_PATH = os.path.join(SCRIPT_DIR, "..", "Lead_Tracker_and_Dashboard.xlsx")
LEADS_SHEET_NAME = "Leads"

COLUMN_ORDER = [
    "Lead ID", "Business Name", "Contact Person", "Phone Number", "City/Area",
    "Category", "Source", "Date Added", "Assigned To", "Status",
    "Follow-up Date", "Deal Value (INR)", "Domain Name", "Payment Received",
    "Site Delivered", "Notes", "Website", "Rating", "Email",
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Scrape Google Maps leads via Apify and append them to the Excel lead tracker."
    )
    parser.add_argument("--category", required=True, help='Business category, e.g. "hardware shop"')
    parser.add_argument("--location", required=True, help='Location, e.g. "Jayanagar, Bangalore"')
    parser.add_argument("--max", type=int, default=100, help="Maximum number of results to scrape (default: 100)")
    parser.add_argument(
        "--excel-path",
        default=os.environ.get("EXCEL_FILE_PATH", DEFAULT_EXCEL_PATH),
        help="Path to the Excel lead tracker (default: %(default)s, override with EXCEL_FILE_PATH in .env)",
    )
    return parser.parse_args()


def get_api_token():
    load_dotenv()
    token = os.environ.get("APIFY_API_TOKEN")
    if not token:
        sys.exit(
            "APIFY_API_TOKEN not found. Add it to lead-scraper/.env as APIFY_API_TOKEN=your_token "
            "before running this script."
        )
    return token


# --- Apify -------------------------------------------------------------

def start_run(token, category, location, max_results):
    url = f"{APIFY_BASE_URL}/acts/{APIFY_ACTOR_ID}/runs"
    payload = {
        "searchStringsArray": [category],
        "locationQuery": location,
        "maxCrawledPlacesPerSearch": max_results,
        "language": "en",
    }
    resp = requests.post(url, params={"token": token}, json=payload, timeout=30)
    resp.raise_for_status()
    data = resp.json()["data"]
    print(f"Started Apify run {data['id']} (actor {APIFY_ACTOR_ID})")
    return data["id"]


def wait_for_run(token, run_id):
    url = f"{APIFY_BASE_URL}/actor-runs/{run_id}"
    waited = 0
    while waited < MAX_POLL_SECONDS:
        resp = requests.get(url, params={"token": token}, timeout=30)
        resp.raise_for_status()
        data = resp.json()["data"]
        status = data["status"]
        print(f"  run status: {status} ({waited}s elapsed)")
        if status in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
            if status != "SUCCEEDED":
                sys.exit(f"Apify run ended with status {status}, aborting.")
            return data["defaultDatasetId"]
        time.sleep(POLL_INTERVAL_SECONDS)
        waited += POLL_INTERVAL_SECONDS
    sys.exit(f"Apify run did not finish within {MAX_POLL_SECONDS}s, aborting.")


def fetch_dataset_items(token, dataset_id):
    url = f"{APIFY_BASE_URL}/datasets/{dataset_id}/items"
    resp = requests.get(url, params={"token": token, "format": "json"}, timeout=60)
    resp.raise_for_status()
    return resp.json()


# --- Cleaning ------------------------------------------------------------

def normalize_phone(raw):
    """Strip +91 / spaces / dashes and return a plain 10-digit number, or None if invalid."""
    if not raw:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits if len(digits) == 10 else None


def clean_records(raw_items, category):
    """Keep rows with a valid phone number and drop duplicate phones within this batch."""
    seen_phones = set()
    cleaned = []
    for item in raw_items:
        phone = normalize_phone(item.get("phoneUnformatted") or item.get("phone"))
        if not phone or phone in seen_phones:
            continue
        seen_phones.add(phone)
        emails = item.get("emails") or []
        cleaned.append(
            {
                "business_name": item.get("title") or "",
                "phone": phone,
                "category": item.get("categoryName") or category,
                "address": item.get("address") or "",
                "website": item.get("website") or None,
                "rating": item.get("totalScore"),
                "email": emails[0] if emails else None,
            }
        )
    return cleaned


# --- Excel -----------------------------------------------------------------

def open_leads_sheet(path):
    if not os.path.isfile(path):
        sys.exit(f"Excel file not found: {path}")
    wb = load_workbook(path)
    if LEADS_SHEET_NAME not in wb.sheetnames:
        sys.exit(f'"{LEADS_SHEET_NAME}" tab not found in {path}')
    ws = wb[LEADS_SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    missing = [col for col in COLUMN_ORDER if col not in headers]
    if missing:
        sys.exit(f'"{LEADS_SHEET_NAME}" tab is missing expected column(s): {", ".join(missing)}')
    return wb, ws, headers


def existing_phones(ws, headers):
    phone_col = headers.index("Phone Number")
    phones = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        normalized = normalize_phone(row[phone_col])
        if normalized:
            phones.add(normalized)
    return phones


def next_lead_id(ws, headers):
    """Continue numbering from the highest trailing number found in the Lead ID column."""
    id_col = headers.index("Lead ID")
    max_num = 0
    prefix = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = row[id_col]
        if value is None:
            continue
        match = re.match(r"^(.*?)(\d+)$", str(value).strip())
        if match:
            num = int(match.group(2))
            if num >= max_num:
                max_num = num
                prefix = match.group(1)
    return prefix, max_num + 1


def append_leads(ws, records, location, prefix, start_id):
    today = date.today()
    lead_id = start_id
    for rec in records:
        row = {
            "Lead ID": f"{prefix}{lead_id}" if prefix else lead_id,
            "Business Name": rec["business_name"],
            "Contact Person": None,
            "Phone Number": rec["phone"],
            "City/Area": location,
            "Category": rec["category"],
            "Source": "Google Maps",
            "Date Added": today,
            "Assigned To": None,
            "Status": "New",
            "Follow-up Date": None,
            "Deal Value (INR)": None,
            "Domain Name": None,
            "Payment Received": None,
            "Site Delivered": None,
            "Notes": None,
            "Website": rec["website"],
            "Rating": rec["rating"],
            "Email": rec["email"],
        }
        ws.append([row[col] for col in COLUMN_ORDER])
        lead_id += 1


# --- Main --------------------------------------------------------------

def main():
    args = parse_args()
    excel_path = os.path.abspath(args.excel_path)

    token = get_api_token()

    print(f'Searching Google Maps for "{args.category}" in "{args.location}" (max {args.max})...')
    run_id = start_run(token, args.category, args.location, args.max)
    dataset_id = wait_for_run(token, run_id)
    raw_items = fetch_dataset_items(token, dataset_id)
    scraped_count = len(raw_items)
    print(f"Scraped {scraped_count} places from Google Maps.")

    cleaned = clean_records(raw_items, args.category)
    dropped_in_batch = scraped_count - len(cleaned)

    wb, ws, headers = open_leads_sheet(excel_path)
    known_phones = existing_phones(ws, headers)

    new_records = [r for r in cleaned if r["phone"] not in known_phones]
    duplicate_existing = len(cleaned) - len(new_records)

    prefix, start_id = next_lead_id(ws, headers)
    append_leads(ws, new_records, args.location, prefix, start_id)

    wb.save(excel_path)

    print("\nSummary")
    print("-------")
    print(f"Scraped from Google Maps:        {scraped_count}")
    print(f"Dropped (no/duplicate phone):    {dropped_in_batch}")
    print(f"Skipped (already in tracker):    {duplicate_existing}")
    print(f"New leads added to '{LEADS_SHEET_NAME}':     {len(new_records)}")
    print(f"Saved to: {excel_path}")


if __name__ == "__main__":
    main()
