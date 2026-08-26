#!/usr/bin/env python3
"""One-time migration: reads the "Leads" tab of Lead_Tracker_and_Dashboard.xlsx
and inserts every row into the Supabase "leads" table.

Safe to re-run: rows are matched by phone_number (which is UNIQUE in the
Supabase schema). Any phone number already in Supabase is left untouched
(not overwritten) — so running this again after someone has edited a lead
in the dashboard won't clobber their changes with stale Excel data.

Usage:
    python3 migrate_excel_to_supabase.py --excel-path /path/to/Lead_Tracker_and_Dashboard.xlsx [--dry-run]

Reads SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY from a .env file in this
script's directory (or the environment). The service_role key is required
because it's the only key allowed to bypass Row Level Security for a
server-side bulk write — never put the service_role key in the Next.js
app's client-side code.
"""

import argparse
import os
import re
import sys
from datetime import date, datetime

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LEADS_SHEET_NAME = "Leads"

# Excel column header -> Supabase column name. Columns not listed here
# (Lead ID, Domain Name is kept, Payment Received / Site Delivered are
# converted below) are handled explicitly in build_row().
EXCEL_TO_DB = {
    "Business Name": "business_name",
    "Contact Person": "contact_person",
    "Phone Number": "phone_number",
    "City/Area": "city_area",
    "Category": "category",
    "Source": "source",
    "Assigned To": "assigned_to",
    "Status": "status",
    "Follow-up Date": "follow_up_date",
    "Deal Value (INR)": "deal_value",
    "Domain Name": "domain_name",
    "Notes": "notes",
    "Website": "website",
    "Rating": "rating",
    "Email": "email",
}

VALID_STATUSES = {"New", "Contacted", "Follow-up", "Interested", "Not Interested", "Closed-Won", "Closed-Lost"}
VALID_ASSIGNED_TO = {"Me", "Friend 1", "Friend 2"}


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--excel-path", required=True, help="Path to Lead_Tracker_and_Dashboard.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be inserted, don't call Supabase")
    return parser.parse_args()


def get_supabase_config():
    load_dotenv(os.path.join(SCRIPT_DIR, "..", ".env.local"))
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit(
            "SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set (in ../.env.local or the environment)."
        )
    return url, key


def to_iso_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return str(value).strip() or None


def to_number(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_phone(raw):
    if not raw:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits if len(digits) == 10 else str(raw).strip()


def build_row(headers, row_cells):
    values = dict(zip(headers, row_cells))
    row = {}

    for excel_col, db_col in EXCEL_TO_DB.items():
        row[db_col] = values.get(excel_col)

    row["business_name"] = (row.get("business_name") or "").strip()
    row["phone_number"] = normalize_phone(row.get("phone_number"))

    status = row.get("status")
    row["status"] = status if status in VALID_STATUSES else "New"

    assigned_to = row.get("assigned_to")
    row["assigned_to"] = assigned_to if assigned_to in VALID_ASSIGNED_TO else None

    row["follow_up_date"] = to_iso_date(row.get("follow_up_date"))
    row["deal_value"] = to_number(row.get("deal_value"))
    row["rating"] = to_number(row.get("rating"))

    date_added = to_iso_date(values.get("Date Added"))
    if date_added:
        row["date_added"] = date_added

    row["payment_received"] = str(values.get("Payment Received") or "").strip().lower() == "yes"
    row["site_delivered"] = str(values.get("Site Delivered") or "").strip().lower() == "yes"

    for key in ("contact_person", "city_area", "category", "source", "domain_name", "notes", "website", "email"):
        if row.get(key) is not None:
            row[key] = str(row[key]).strip() or None

    return row


def read_excel_rows(excel_path):
    if not os.path.isfile(excel_path):
        sys.exit(f"Excel file not found: {excel_path}")
    wb = load_workbook(excel_path)
    if LEADS_SHEET_NAME not in wb.sheetnames:
        sys.exit(f'"{LEADS_SHEET_NAME}" tab not found in {excel_path}')
    ws = wb[LEADS_SHEET_NAME]
    headers = [c.value for c in ws[1]]

    rows = []
    skipped = []
    for row_number in range(2, ws.max_row + 1):
        cells = [c.value for c in ws[row_number]]
        if all(v is None for v in cells):
            continue
        row = build_row(headers, cells)
        if not row["business_name"] or not row["phone_number"] or len(row["phone_number"]) != 10:
            skipped.append((row_number, row.get("business_name"), row.get("phone_number")))
            continue
        rows.append(row)
    return rows, skipped


def insert_new_rows(supabase_url, service_key, rows):
    """Insert rows, skipping (not overwriting) any phone_number that already
    exists — so re-running this after someone has edited a lead in the
    dashboard never clobbers their changes with stale Excel data."""
    if not rows:
        return 0
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/leads"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates,return=representation",
    }
    resp = requests.post(f"{endpoint}?on_conflict=phone_number", headers=headers, json=rows, timeout=60)
    resp.raise_for_status()
    return len(resp.json())


def main():
    args = parse_args()
    rows, skipped = read_excel_rows(args.excel_path)

    print(f"Read {len(rows) + len(skipped)} data row(s) from {args.excel_path}")
    print(f"  {len(rows)} row(s) ready to import")
    print(f"  {len(skipped)} row(s) skipped (missing business name or a valid 10-digit phone number)")
    for row_number, name, phone in skipped:
        print(f"    - row {row_number}: business_name={name!r} phone_number={phone!r}")

    if args.dry_run:
        print("\n--dry-run set, not calling Supabase. Example row that would be sent:")
        if rows:
            print(rows[0])
        return

    if not rows:
        print("\nNothing to import.")
        return

    url, key = get_supabase_config()
    inserted = insert_new_rows(url, key, rows)
    already_present = len(rows) - inserted
    print(f"\nMigrated {inserted} new row(s) into Supabase.")
    if already_present:
        print(f"Skipped {already_present} row(s) already present in Supabase (matched by phone_number) — left untouched.")


if __name__ == "__main__":
    main()
