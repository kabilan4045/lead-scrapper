"""Read/write access to the "Leads" tab of the Excel lead tracker.

Only ever touches the Leads sheet — the Dashboard sheet (and its formulas
and formatting) is never opened for writing by this module.
"""

import threading
from datetime import date, datetime

from openpyxl import load_workbook

LEADS_SHEET_NAME = "Leads"

STATUS_OPTIONS = ["New", "Contacted", "Follow-up", "Interested", "Closed-Won", "Closed-Lost"]
ASSIGNED_TO_OPTIONS = ["Me", "Chetan", "Nandhu"]

# Internal field name -> Excel column header. Order here is display order.
DISPLAY_COLUMNS = {
    "lead_id": "Lead ID",
    "business_name": "Business Name",
    "phone": "Phone Number",
    "city_area": "City/Area",
    "category": "Category",
    "assigned_to": "Assigned To",
    "status": "Status",
    "follow_up_date": "Follow-up Date",
    "deal_value": "Deal Value (INR)",
    "website": "Website",
    "notes": "Notes",
}

# Fields the app is allowed to write. Everything else is read-only from here.
EDITABLE_COLUMNS = {
    "status": "Status",
    "assigned_to": "Assigned To",
    "follow_up_date": "Follow-up Date",
    "notes": "Notes",
}

_lock = threading.Lock()


class ExcelIOError(Exception):
    """Raised when the workbook can't be read/saved (e.g. open in Excel elsewhere)."""


def _open_leads_ws(path):
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        raise ExcelIOError(f"Excel file not found: {path}")
    except Exception as exc:  # openpyxl raises various errors for locked/corrupt files
        raise ExcelIOError(f"Could not open the Excel file — is it open in another program? ({exc})")

    if LEADS_SHEET_NAME not in wb.sheetnames:
        raise ExcelIOError(f'"{LEADS_SHEET_NAME}" tab not found in {path}')

    ws = wb[LEADS_SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    header_col = {}
    for name in set(DISPLAY_COLUMNS.values()):
        if name not in headers:
            raise ExcelIOError(f'"{LEADS_SHEET_NAME}" tab is missing expected column: {name}')
        header_col[name] = headers.index(name) + 1  # 1-based for openpyxl cell access

    return wb, ws, header_col


def _cell_to_json(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return value


def _read_row(ws, header_col, row_number):
    lead = {"row": row_number}
    for field, col_name in DISPLAY_COLUMNS.items():
        lead[field] = _cell_to_json(ws.cell(row=row_number, column=header_col[col_name]).value)
    return lead


def read_leads(path):
    """Return every lead row from the Leads tab as a list of plain dicts."""
    with _lock:
        wb, ws, header_col = _open_leads_ws(path)
        leads = []
        for row_number in range(2, ws.max_row + 1):
            lead_id = ws.cell(row=row_number, column=header_col["Lead ID"]).value
            if lead_id is None:
                continue
            leads.append(_read_row(ws, header_col, row_number))
        return leads


def _parse_follow_up_date(value):
    if not value:
        return None
    if isinstance(value, str):
        return datetime.strptime(value, "%Y-%m-%d").date()
    raise ExcelIOError(f"Unrecognized follow_up_date value: {value!r}")


def update_lead(path, lead_id, fields):
    """Update one or more editable fields on the row matching lead_id.

    `fields` is a dict of editable-field-name -> new value (as received from the
    browser, i.e. plain strings). Returns the row's fresh state as a dict.
    """
    unknown = set(fields) - set(EDITABLE_COLUMNS)
    if unknown:
        raise ExcelIOError(f"These fields can't be edited: {', '.join(sorted(unknown))}")

    with _lock:
        wb, ws, header_col = _open_leads_ws(path)

        target_row = None
        for row_number in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_number, column=header_col["Lead ID"]).value
            if cell_value is not None and str(cell_value).strip() == str(lead_id).strip():
                target_row = row_number
                break

        if target_row is None:
            raise ExcelIOError(f"Lead ID {lead_id} not found")

        for field, raw_value in fields.items():
            col_name = EDITABLE_COLUMNS[field]
            if field == "follow_up_date":
                value = _parse_follow_up_date(raw_value)
            elif field == "status":
                if raw_value not in STATUS_OPTIONS:
                    raise ExcelIOError(f"Invalid status: {raw_value!r}")
                value = raw_value
            elif field == "assigned_to":
                if raw_value and raw_value not in ASSIGNED_TO_OPTIONS:
                    raise ExcelIOError(f"Invalid assigned_to: {raw_value!r}")
                value = raw_value or None
            else:  # notes — free text
                value = raw_value or None
            # ws.cell(..., value=value) treats value=None as "leave unchanged" (it shares
            # None with the parameter's default), so clearing a field would silently no-op.
            # Assigning .value directly does not have that ambiguity.
            ws.cell(row=target_row, column=header_col[col_name]).value = value

        try:
            wb.save(path)
        except Exception as exc:
            raise ExcelIOError(f"Could not save the Excel file — is it open in another program? ({exc})")

        return _read_row(ws, header_col, target_row)
