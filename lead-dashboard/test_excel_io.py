"""Exercises excel_io.py against a scratch copy of the backup workbook.

Run this any time you want to sanity-check the read/write logic without
touching the real Lead_Tracker_and_Dashboard.xlsx. It never opens the real
file or the pristine backup directly — it copies the backup to a throwaway
file first.
"""

import os
import shutil
import sys
from datetime import date

from openpyxl import load_workbook

import excel_io

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BACKUP_PATH = os.path.join(SCRIPT_DIR, "backups", "Lead_Tracker_and_Dashboard_backup_2026-08-26.xlsx")
SCRATCH_PATH = os.path.join(SCRIPT_DIR, "backups", "_test_scratch.xlsx")

failures = []


def check(label, condition):
    status = "PASS" if condition else "FAIL"
    print(f"[{status}] {label}")
    if not condition:
        failures.append(label)


def seed_scratch_workbook(target_path=SCRATCH_PATH):
    shutil.copy(BACKUP_PATH, target_path)
    wb = load_workbook(target_path)
    ws = wb["Leads"]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    def row(lead_id, name, phone, city, category, assigned_to, status, follow_up, deal_value, website, notes):
        r = ws.max_row + 1
        ws.cell(row=r, column=col["Lead ID"], value=lead_id)
        ws.cell(row=r, column=col["Business Name"], value=name)
        ws.cell(row=r, column=col["Phone Number"], value=phone)
        ws.cell(row=r, column=col["City/Area"], value=city)
        ws.cell(row=r, column=col["Category"], value=category)
        ws.cell(row=r, column=col["Source"], value="Google Maps")
        ws.cell(row=r, column=col["Date Added"], value=date(2026, 8, 20))
        ws.cell(row=r, column=col["Assigned To"], value=assigned_to)
        ws.cell(row=r, column=col["Status"], value=status)
        ws.cell(row=r, column=col["Follow-up Date"], value=follow_up)
        ws.cell(row=r, column=col["Deal Value (INR)"], value=deal_value)
        ws.cell(row=r, column=col["Website"], value=website)
        ws.cell(row=r, column=col["Notes"], value=notes)

    row(1, "A Hardware", "9876543210", "Jayanagar, Bangalore", "Hardware store", "Me", "New", None, None, None, None)
    row(2, "B Hardware", "9876543211", "Jayanagar, Bangalore", "Hardware store", "Friend 1", "Contacted", date(2026, 9, 1), 15000, "https://bhardware.example.com", "Called once")
    row(3, "C Hardware", "9876543212", "Koramangala, Bangalore", "Hardware store", None, "Closed-Won", None, 22000, None, None)

    wb.save(target_path)


def main():
    if not os.path.isfile(BACKUP_PATH):
        sys.exit(f"Backup not found at {BACKUP_PATH} — run the backup step first.")

    seed_scratch_workbook()

    # --- read_leads ---
    leads = excel_io.read_leads(SCRATCH_PATH)
    check("read_leads returns 3 seeded rows", len(leads) == 3)
    check("row fields include all display columns", set(leads[0].keys()) >= set(excel_io.DISPLAY_COLUMNS) | {"row"})
    check("follow_up_date serializes to ISO string", leads[1]["follow_up_date"] == "2026-09-01")
    check("blank follow_up_date reads as None", leads[0]["follow_up_date"] is None)
    check("deal_value reads as a number", leads[1]["deal_value"] == 15000)
    check("website reads through", leads[1]["website"] == "https://bhardware.example.com")

    # --- update_lead: happy path on each editable field ---
    updated = excel_io.update_lead(SCRATCH_PATH, 1, {"status": "Interested"})
    check("update status persists", updated["status"] == "Interested")

    updated = excel_io.update_lead(SCRATCH_PATH, 1, {"assigned_to": "Friend 2"})
    check("update assigned_to persists", updated["assigned_to"] == "Friend 2")

    updated = excel_io.update_lead(SCRATCH_PATH, 1, {"follow_up_date": "2026-09-15"})
    check("update follow_up_date persists", updated["follow_up_date"] == "2026-09-15")

    updated = excel_io.update_lead(SCRATCH_PATH, 1, {"notes": "Left voicemail"})
    check("update notes persists", updated["notes"] == "Left voicemail")

    # Multiple fields in one call
    updated = excel_io.update_lead(SCRATCH_PATH, 2, {"status": "Closed-Lost", "notes": "Went with a competitor"})
    check("multi-field update: status", updated["status"] == "Closed-Lost")
    check("multi-field update: notes", updated["notes"] == "Went with a competitor")

    # Re-read from disk to make sure the save actually landed, not just in-memory
    reloaded = {r["lead_id"]: r for r in excel_io.read_leads(SCRATCH_PATH)}
    check("row 1 edits persisted to disk", reloaded[1]["status"] == "Interested" and reloaded[1]["notes"] == "Left voicemail")
    check("row 2 edits persisted to disk", reloaded[2]["status"] == "Closed-Lost")
    check("row 3 (untouched) still Closed-Won", reloaded[3]["status"] == "Closed-Won")
    check("editing row 1 didn't touch row 2's business name", reloaded[2]["business_name"] == "B Hardware")

    # --- error handling ---
    try:
        excel_io.update_lead(SCRATCH_PATH, 999, {"status": "New"})
        check("unknown lead_id raises", False)
    except excel_io.ExcelIOError:
        check("unknown lead_id raises", True)

    try:
        excel_io.update_lead(SCRATCH_PATH, 1, {"status": "Not A Real Status"})
        check("invalid status value raises", False)
    except excel_io.ExcelIOError:
        check("invalid status value raises", True)

    try:
        excel_io.update_lead(SCRATCH_PATH, 1, {"business_name": "Hacked Name"})
        check("writing a non-editable field raises", False)
    except excel_io.ExcelIOError:
        check("writing a non-editable field raises", True)

    reloaded = {r["lead_id"]: r for r in excel_io.read_leads(SCRATCH_PATH)}
    check("rejected business_name write left data untouched", reloaded[1]["business_name"] == "A Hardware")

    # --- Dashboard tab must be completely untouched ---
    wb = load_workbook(SCRATCH_PATH)
    check('"Dashboard" sheet still present', "Dashboard" in wb.sheetnames)
    dash = wb["Dashboard"]
    dash_values = [dash.cell(row=r, column=2).value for r in range(3, 11)]
    check("Dashboard formulas are still formula strings (unmodified)", all(isinstance(v, str) and v.startswith("=") for v in dash_values))
    check("Dashboard title cell untouched", dash["A1"].value == "Lead Tracker Dashboard")

    os.remove(SCRATCH_PATH)

    print()
    if failures:
        print(f"{len(failures)} check(s) FAILED:")
        for f in failures:
            print(" -", f)
        sys.exit(1)
    else:
        print("All checks passed.")


if __name__ == "__main__":
    main()
