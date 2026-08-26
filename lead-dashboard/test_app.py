"""HTTP-level integration test for the Flask app, run against a scratch copy
of the backup workbook (never the real file, never even the pristine backup).
"""

import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SCRATCH_PATH = os.path.join(SCRIPT_DIR, "backups", "_test_app_scratch.xlsx")

sys.path.insert(0, SCRIPT_DIR)
from test_excel_io import seed_scratch_workbook  # noqa: E402

seed_scratch_workbook(SCRATCH_PATH)
os.environ["EXCEL_FILE_PATH"] = SCRATCH_PATH  # must be set before importing app

import app  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

failures = []


def check(label, condition):
    status = "PASS" if condition else "FAIL"
    print(f"[{status}] {label}")
    if not condition:
        failures.append(label)


def main():
    client = app.app.test_client()

    resp = client.get("/")
    check("GET / returns 200", resp.status_code == 200)
    check("GET / renders the page title", b"Lead Dashboard" in resp.data)

    resp = client.get("/api/leads")
    check("GET /api/leads returns 200", resp.status_code == 200)
    leads = resp.get_json()
    check("GET /api/leads returns 3 seeded rows", len(leads) == 3)

    resp = client.patch("/api/leads/1", json={"status": "Interested", "notes": "Left voicemail"})
    check("PATCH valid fields returns 200", resp.status_code == 200)
    body = resp.get_json()
    check("PATCH response reflects new status", body["status"] == "Interested")
    check("PATCH response reflects new notes", body["notes"] == "Left voicemail")

    resp = client.get("/api/leads")
    leads = {l["lead_id"]: l for l in resp.get_json()}
    check("PATCH persisted across a fresh GET", leads[1]["status"] == "Interested" and leads[1]["notes"] == "Left voicemail")

    resp = client.patch("/api/leads/1", json={"follow_up_date": "2026-10-01"})
    check("PATCH follow_up_date returns 200", resp.status_code == 200)
    check("PATCH follow_up_date reflected", resp.get_json()["follow_up_date"] == "2026-10-01")

    resp = client.patch("/api/leads/1", json={"follow_up_date": ""})
    check("Clearing follow_up_date returns 200", resp.status_code == 200)
    check("Cleared follow_up_date reads back as null", resp.get_json()["follow_up_date"] is None)

    resp = client.patch("/api/leads/999", json={"status": "New"})
    check("PATCH unknown lead_id returns 400", resp.status_code == 400)

    resp = client.patch("/api/leads/1", json={"status": "Not A Real Status"})
    check("PATCH invalid status returns 400", resp.status_code == 400)

    resp = client.patch("/api/leads/1", json={"business_name": "Hacked"})
    check("PATCH non-editable field returns 400", resp.status_code == 400)

    resp = client.patch("/api/leads/1", json={})
    check("PATCH empty body returns 400", resp.status_code == 400)

    resp = client.get("/api/leads")
    leads = {l["lead_id"]: l for l in resp.get_json()}
    check("Rejected writes left row 1 business_name untouched", leads[1]["business_name"] == "A Hardware")
    check("Row 2 was never touched by any of this", leads[2]["status"] == "Contacted")

    wb = load_workbook(SCRATCH_PATH)
    check('"Dashboard" sheet still present after HTTP traffic', "Dashboard" in wb.sheetnames)
    dash = wb["Dashboard"]
    dash_values = [dash.cell(row=r, column=2).value for r in range(3, 11)]
    check("Dashboard formulas untouched after HTTP traffic", all(isinstance(v, str) and v.startswith("=") for v in dash_values))

    os.remove(SCRATCH_PATH)

    print()
    if failures:
        print(f"{len(failures)} check(s) FAILED:")
        for f in failures:
            print(" -", f)
        sys.exit(1)
    else:
        print("All checks passed.")


if __name__ == "__main__":
    main()
